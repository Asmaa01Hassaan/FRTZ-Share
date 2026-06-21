# -*- coding: utf-8 -*-
import logging

from odoo import fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class SubscriptionLinkInvoiceWizard(models.TransientModel):
    """Link existing (historical/migrated) customer invoices to a subscription,
    and/or import historical invoices from an Excel file."""
    _name = 'subscription.link.invoice.wizard'
    _description = 'Link Existing Invoices to Subscription'

    order_id = fields.Many2one(
        'sale.order', string='Subscription', required=True, readonly=True, ondelete='cascade')
    partner_id = fields.Many2one(related='order_id.partner_id', readonly=True)
    commercial_partner_id = fields.Many2one(
        related='order_id.partner_id.commercial_partner_id', readonly=True)
    company_id = fields.Many2one(related='order_id.company_id', readonly=True)
    set_invoice_origin = fields.Boolean(
        string="Set 'Source Document'", default=True,
        help="Stamp the subscription reference as the invoice's Source Document "
             "when linking (only if currently empty).")
    move_ids = fields.Many2many(
        'account.move', 'subscription_link_wizard_move_rel', 'wizard_id', 'move_id',
        string='Invoices to Link')

    # Excel import of historical invoices
    import_file = fields.Binary(string='Excel File', attachment=False)
    import_filename = fields.Char(string='File Name')
    post_invoices = fields.Boolean(
        string='Post & Number Invoices', default=True,
        help="Post the imported invoices so Odoo assigns each one the next "
             "invoice number in its journal sequence. Leave unticked to keep "
             "them as drafts.")

    def action_link(self):
        self.ensure_one()
        self.order_id._check_is_subscription()
        moves = self.move_ids
        if not moves:
            raise UserError(_("Select at least one invoice to link."))
        blocked = moves.filtered('subscription_order_id')
        linkable = moves - blocked
        for move in linkable:
            move_vals = {
                'subscription_order_id': self.order_id.id,
                'subscription_linked_manually': True,
            }
            if self.set_invoice_origin and not move.invoice_origin:
                move_vals['invoice_origin'] = self.order_id.name
            move.sudo().write(move_vals)
        if linkable:
            self.order_id.message_post(body=_(
                "Linked %(n)s existing invoice(s): %(refs)s",
                n=len(linkable), refs=", ".join(linkable.mapped('display_name'))))
        if blocked:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _("Partially linked"),
                    'message': _("%(linked)s linked, %(skipped)s already linked elsewhere.",
                                 linked=len(linkable), skipped=len(blocked)),
                    'type': 'warning',
                    'sticky': False,
                },
            }
        return self.order_id.action_view_subscription_invoices()

    # ------------------------------------------------------------------
    # Downloadable Excel template
    # ------------------------------------------------------------------
    _TEMPLATE_HEADERS = [
        'Reference', 'Invoice Date', 'Product', 'Description',
        'Quantity', 'Amount', 'Period Start', 'Period End',
    ]

    def _build_template_xlsx(self):
        from openpyxl import Workbook
        import io
        wb = Workbook()
        ws = wb.active
        ws.title = 'Invoices'
        ws.append(self._TEMPLATE_HEADERS)
        ws.append(['INV-0001', '2025-01-31', 'Monthly Package',
                   'Subscription - January', 1, 100, '2025-01-01', '2025-01-31'])
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 22
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def action_download_template(self):
        self.ensure_one()
        import base64
        attachment = self.env['ir.attachment'].create({
            'name': 'subscription_invoices_template.xlsx',
            'datas': base64.b64encode(self._build_template_xlsx()),
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Import historical invoices from an Excel file
    # ------------------------------------------------------------------
    @staticmethod
    def _xls_header_index(headers, *names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    def action_import_invoices_excel(self):
        """Read an .xlsx file (one invoice per row) and create historical customer
        invoices linked to this subscription. Expected header row columns:
        Reference, Invoice Date, Description, Quantity, Amount, Period Start, Period End.
        """
        self.ensure_one()
        self.order_id._check_is_subscription()
        if not self.import_file:
            raise UserError(_("Please choose an Excel (.xlsx) file first."))
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("The 'openpyxl' library is required to read Excel files."))
        import base64
        import io
        from datetime import datetime, date as date_cls

        try:
            wb = load_workbook(io.BytesIO(base64.b64decode(self.import_file)),
                               data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(_("Could not read the Excel file: %s") % exc)
        rows = list(wb.active.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_("The file has no data rows (a header row + at least one line are expected)."))

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
        idx_ref = self._xls_header_index(headers, 'reference', 'number', 'ref', 'invoice number')
        idx_date = self._xls_header_index(headers, 'invoice date', 'date')
        idx_product = self._xls_header_index(headers, 'product', 'product name', 'item')
        idx_desc = self._xls_header_index(headers, 'description', 'desc', 'label')
        idx_qty = self._xls_header_index(headers, 'quantity', 'qty')
        idx_amount = self._xls_header_index(headers, 'amount', 'unit price', 'price', 'total')
        idx_pstart = self._xls_header_index(headers, 'period start', 'start')
        idx_pend = self._xls_header_index(headers, 'period end', 'end')
        if idx_amount is None:
            raise UserError(_("The file must contain an 'Amount' column."))

        def cell(row, idx):
            return row[idx] if (idx is not None and idx < len(row)) else None

        Product = self.env['product.product']
        new_product_ids = []

        def find_or_create_product(value):
            """Match the Excel product to an Odoo product (by internal reference or
            name, case-insensitive); create it as a service product if missing so
            the invoice line always carries the same product as the file."""
            if not value:
                return Product
            text = str(value).strip()
            product = Product.search(
                ['|', ('default_code', '=ilike', text), ('name', '=ilike', text)], limit=1)
            if not product:
                product = Product.search([('name', 'ilike', text)], limit=1)
            if not product:
                product = Product.sudo().create({
                    'name': text, 'type': 'service', 'sale_ok': True})
                new_product_ids.append(product.id)
            return product

        def to_date(value):
            if not value:
                return False
            if isinstance(value, (datetime, date_cls)):
                return fields.Date.to_date(value)
            try:
                return fields.Date.to_date(str(value))
            except Exception:
                return False

        def to_float(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0.0

        partner = self.order_id._get_subscription_invoice_partner()
        AccountMove = self.env['account.move'].sudo()
        created = self.env['account.move']
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            amount = to_float(cell(row, idx_amount))
            qty = to_float(cell(row, idx_qty)) or 1.0
            product = find_or_create_product(cell(row, idx_product))
            desc_raw = cell(row, idx_desc)
            desc = str(desc_raw).strip() if desc_raw else False
            # The product line shows the description; if no description is given,
            # fall back to the product name (or a generic label).
            line_name = desc or (product.display_name if product else _("Subscription invoice"))
            # Historical import: the Amount column is the final line amount, so no
            # taxes are auto-applied (this also avoids the product's default taxes
            # clashing with the company fiscal country).
            invoice_line = {
                'name': line_name, 'quantity': qty, 'price_unit': amount,
                'tax_ids': [(6, 0, [])],
            }
            if product:
                invoice_line['product_id'] = product.id
            move_vals = {
                'move_type': 'out_invoice',
                'partner_id': partner.id,
                'invoice_date': to_date(cell(row, idx_date)) or False,
                # 'Reference' column maps to the invoice Bill Reference (ref).
                'ref': (str(cell(row, idx_ref)).strip() if cell(row, idx_ref) else False),
                'invoice_origin': self.order_id.name,
                'subscription_order_id': self.order_id.id,
                'subscription_linked_manually': True,
                'subscription_period_start': to_date(cell(row, idx_pstart)),
                'subscription_period_end': to_date(cell(row, idx_pend)),
                'invoice_line_ids': [(0, 0, invoice_line)],
            }
            move = AccountMove.create(move_vals)
            # Setting product_id makes the line re-price from the product; re-assert
            # the amount from the file as a manual override so it sticks.
            if product:
                line = move.invoice_line_ids[:1]
                if line and line.price_unit != amount:
                    line.with_context(check_move_validity=False).price_unit = amount
            created |= move
        if not created:
            raise UserError(_("No invoice rows were found in the file."))

        # Post the imported invoices so each gets the next Odoo invoice number.
        posted = self.env['account.move']
        failed = []
        if self.post_invoices:
            for move in created:
                try:
                    move.action_post()
                    posted |= move
                except Exception as exc:  # noqa: BLE001 - one bad row must not abort the batch
                    failed.append(move.ref or _("(row %s)") % move.id)
                    _logger.warning("Imported invoice could not be posted (%s): %s",
                                    move.ref or move.id, exc)
        message = _("Imported %s invoice(s) from Excel.") % len(created)
        if new_product_ids:
            message += " " + _("%s new product(s) created.") % len(set(new_product_ids))
        if self.post_invoices:
            message += " " + _("%s posted and numbered.") % len(posted)
            if failed:
                message += " " + _(
                    "%(n)s kept as draft (could not be posted - check taxes/accounts): %(refs)s",
                    n=len(failed), refs=", ".join(failed))
        self.order_id.message_post(body=message)
        return self.order_id.action_view_subscription_invoices()
